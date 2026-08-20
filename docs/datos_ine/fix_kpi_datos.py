"""Fix circular refs / #VALUE! in KPI_Datos sheet."""
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

path = r"e:\Users\gracobjo\Documents\proyectosMios\culebra\Modelo_Cuenta_Resultados_Marketplace_5_anos.xlsx"
wb = load_workbook(path)
ws = wb["KPI_Datos"]

money = "#,##0.00"
pct = "0.00%"
header_fill = PatternFill("solid", fgColor="2F5233")
header_font = Font(bold=True, color="FFFFFF")


def style_header(cell, value):
    cell.value = value
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)


# Restore Table A headers D–H (were overwritten by KPI cards)
for i, h in enumerate(
    ["Gastos", "Neto", "Margen neto", "Take rate", "Cobertura gastos"], 4
):
    style_header(ws.cell(4, i), h)

pyg = "'PyG Conservador'"
for i in range(1, 6):
    r = 4 + i  # rows 5..9
    # A–C already correct: Año / GMV / Ingresos
    ws.cell(r, 4).value = f"=SUMIF({pyg}!$A:$A,A{r},{pyg}!$Q:$Q)"  # Gastos
    ws.cell(r, 5).value = f"=SUMIF({pyg}!$A:$A,A{r},{pyg}!$T:$T)"  # Neto
    ws.cell(r, 6).value = f"=IF(C{r}=0,0,E{r}/C{r})"  # Margen = Neto/Ingresos
    ws.cell(r, 7).value = f"=IF(B{r}=0,0,C{r}/B{r})"  # Take rate = Ingresos/GMV
    ws.cell(r, 8).value = f"=IF(D{r}=0,0,C{r}/D{r})"  # Cobertura = Ingresos/Gastos
    for c in range(4, 6):
        ws.cell(r, c).number_format = money
    for c in range(6, 9):
        ws.cell(r, c).number_format = pct

# Clear leftover card cells that sat on top of Table A (D10:H13)
for r in range(10, 14):
    for c in range(4, 9):
        cell = ws.cell(r, c)
        cell.value = None
        cell.number_format = "General"
        cell.fill = PatternFill()
        cell.font = Font()

# Clear old "F. Tarjetas" title in D3 if present
if ws["D3"].value and "Tarjetas" in str(ws["D3"].value):
    ws["D3"].value = None
    ws["D3"].font = Font()

# Section F cards in J–K (no overlap with annual table)
ws["J3"] = "F. Tarjetas KPI (Conservador)"
ws["J3"].font = Font(bold=True)
style_header(ws.cell(4, 10), "KPI")
style_header(ws.cell(4, 11), "Valor")

cards = [
    (5, "GMV Y3", "=B7", money),
    (6, "Ingresos Y3", "=C7", money),
    (7, "Neto Y3", "=E7", money),  # E7 is now Neto SUMIF — no circular
    (8, "Margen neto Y3", "=F7", pct),
    (9, "Take rate", "=G7", pct),
    (10, "Cobertura gastos Y3", "=H7", pct),
    (11, "GMV equilibrio mes", 9700, money),
    (12, "Vendedores obj. Y3", 15, "0"),
    (13, "Comisión marketplace", 0.15, pct),
]
for r, label, val, fmt in cards:
    ws.cell(r, 10).value = label
    ws.cell(r, 11).value = val
    ws.cell(r, 11).number_format = fmt

ws.column_dimensions["J"].width = 24
ws.column_dimensions["K"].width = 14

ws["A10"] = (
    "Nota: D–H = Gastos/Neto/ratios de la evolución anual. "
    "Tarjetas KPI en J–K (sin solape ni referencias circulares)."
)
ws["A10"].font = Font(italic=True, size=9, color="666666")

wb.save(path)
print("OK: KPI_Datos corregido")
for coord in ["B7", "C7", "D7", "E7", "F7", "G7", "H7", "J7", "K7", "H5"]:
    print(f"  {coord}: {ws[coord].value}")
