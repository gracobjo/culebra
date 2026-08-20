# -*- coding: utf-8 -*-
"""
Poblar Modelo_Cuenta_Resultados_Marketplace_5_anos.xlsx
- Conservador (prudente): PyG mensual activa + Y1 suavizado (amort. desde lanzamiento)
- Realista y Optimista: PyG mensual completa
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "Modelo_Cuenta_Resultados_Marketplace_5_anos.xlsx"
OUT = ROOT / "Modelo_Cuenta_Resultados_Marketplace_5_anos.xlsx"

SEASON = [2752, 3851, 5713, 3900, 5539, 5940, 8380, 12151, 5981, 6633, 5731, 5338]
SEASON_W = [x / sum(SEASON) for x in SEASON]
MONTHS = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]
Y1_ACTIVE = {7, 8, 9, 10, 11, 12}
HEADERS = [
    "Año", "Mes", "GMV vendedores", "Ingresos comisión (15%)", "Coste medios pago",
    "Personal", "Transportes/logística", "Marketing", "Hosting/software",
    "Gestoría/contabilidad", "Seguros", "Telefonía/Internet", "Suministros/oficina",
    "Otros corrientes", "Amortizaciones", "Gastos financieros", "Total gastos",
    "Resultado antes impuestos", "Impuesto Sociedades", "Resultado neto",
]

# GMV anual por escenario (conservador recalibrado: recuperar inversión 30k + ≥15% dividendos en 5 años)
# Objetivo: neto acumulado ≥ 30.000 × 1,15 = 34.500 € con volúmenes más modestos que la v1.
GMV = {
    "conservador": {1: 40_000, 2: 140_000, 3: 220_000, 4: 280_000, 5: 360_000},
    "realista": {1: 75_000, 2: 260_000, 3: 400_000, 4: 520_000, 5: 680_000},
    "optimista": {1: 110_000, 2: 400_000, 3: 620_000, 4: 800_000, 5: 1_050_000},
}

COMISION = 0.15
IS = 0.15
PCT_PAGO = 0.015
AMORT_ANUAL = 6_000.0  # 30k / 5 años; en Y1 solo H2 (lanzamiento)
INVERSION_REF = 30_000.0
DIVIDENDO_OBJ_PCT = 0.15  # ≥15% sobre la inversión en el horizonte de 5 años


def money_fmt(cell):
    cell.number_format = "#,##0.00"


def pct_fmt(cell):
    cell.number_format = "0.00%"


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2F5233")
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")


def gmv_month(scenario: str, year: int, month: int) -> float:
    annual = GMV[scenario][year]
    if year == 1:
        if month not in Y1_ACTIVE:
            return 0.0
        w = sum(SEASON_W[m - 1] for m in Y1_ACTIVE)
        return annual * (SEASON_W[month - 1] / w)
    return annual * SEASON_W[month - 1]


def opex(scenario: str, year: int, month: int) -> dict[str, float]:
    """Costes mensuales (sin medios de pago ni amortización)."""
    launched = not (year == 1 and month not in Y1_ACTIVE)

    # Conservador: opex contenido alineado a GMV más modesto (sin marketing agresivo)
    if scenario == "conservador":
        if year == 1 and not launched:
            return dict(
                personal=0, transportes=0, marketing=0, hosting=80,
                gestoria=120, seguros=30, telefonia=35, suministros=40,
                otros=25, fin=10,
            )
        personal = 0 if year == 1 else 150
        if year == 1:
            marketing, hosting = 350, 180  # ~2.100 mkt + ~1.560 host en H2 (+H1 mínimo)
        elif year == 2:
            marketing, hosting = 400, 220
        else:
            marketing, hosting = 550, 250
        return dict(
            personal=personal, transportes=0, marketing=marketing, hosting=hosting,
            gestoria=120, seguros=30, telefonia=35, suministros=40,
            otros=25, fin=10,
        )

    if scenario == "realista":
        if year == 1 and not launched:
            return dict(
                personal=0, transportes=0, marketing=50, hosting=120,
                gestoria=150, seguros=40, telefonia=45, suministros=55,
                otros=40, fin=15,
            )
        personal = [0, 500, 750, 1000, 1000][year - 1]
        marketing = [700, 1000, 1500, 1667, 1833][year - 1] if launched or year > 1 else 50
        if year == 1 and launched:
            marketing = 700  # ~4.200 en H2 + algo H1
        hosting = [220, 420, 500, 580, 670][year - 1]
        return dict(
            personal=personal, transportes=50, marketing=marketing, hosting=hosting,
            gestoria=180, seguros=50, telefonia=55, suministros=70,
            otros=60, fin=25,
        )

    # optimista
    if year == 1 and not launched:
        return dict(
            personal=0, transportes=0, marketing=100, hosting=150,
            gestoria=180, seguros=50, telefonia=55, suministros=70,
            otros=60, fin=25,
        )
    personal = [0, 1000, 1500, 2000, 2500][year - 1]
    marketing = 1000 if (year == 1 and launched) else [1000, 1500, 2333, 2917, 3333][year - 1]
    hosting = [250, 500, 670, 830, 1000][year - 1]
    return dict(
        personal=personal, transportes=100, marketing=marketing, hosting=hosting,
        gestoria=200, seguros=60, telefonia=60, suministros=80,
        otros=80, fin=40,
    )


def amort(year: int, month: int) -> float:
    """Amortización desde capitalización en lanzamiento (mes 7 año 1)."""
    if year == 1 and month not in Y1_ACTIVE:
        return 0.0
    return AMORT_ANUAL / 12


def row_values(scenario: str, year: int, month: int) -> dict:
    gmv = round(gmv_month(scenario, year, month), 2)
    ing = round(gmv * COMISION, 2)
    pago = round(gmv * PCT_PAGO, 2)
    ox = opex(scenario, year, month)
    am = amort(year, month)
    return {
        "gmv": gmv,
        "ing": ing,
        "pago": pago,
        "personal": ox["personal"],
        "transportes": ox["transportes"],
        "marketing": ox["marketing"],
        "hosting": ox["hosting"],
        "gestoria": ox["gestoria"],
        "seguros": ox["seguros"],
        "telefonia": ox["telefonia"],
        "suministros": ox["suministros"],
        "otros": ox["otros"],
        "amort": am,
        "fin": ox["fin"],
    }


def annual_totals(scenario: str) -> list[dict]:
    out = []
    for year in range(1, 6):
        gmv = ing = gast = 0.0
        for m in range(1, 13):
            v = row_values(scenario, year, m)
            gmv += v["gmv"]
            ing += v["ing"]
            gast += (
                v["pago"] + v["personal"] + v["transportes"] + v["marketing"]
                + v["hosting"] + v["gestoria"] + v["seguros"] + v["telefonia"]
                + v["suministros"] + v["otros"] + v["amort"] + v["fin"]
            )
        rai = ing - gast
        tax = max(0.0, rai * IS)
        out.append({
            "year": year, "gmv": gmv, "ing": ing, "gast": gast,
            "rai": rai, "tax": tax, "neto": rai - tax,
        })
    return out


def write_pyg(ws, scenario: str, title_note: str):
    ws.cell(1, 1).value = title_note
    ws.cell(1, 1).font = Font(bold=True, color="2F5233")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)

    for i, h in enumerate(HEADERS, 1):
        cell = ws.cell(2, i, h)
        style_header(cell)
    ws.row_dimensions[2].height = 30

    r = 3
    for year in range(1, 6):
        for month in range(1, 13):
            v = row_values(scenario, year, month)
            ws.cell(r, 1).value = year
            ws.cell(r, 2).value = MONTHS[month - 1]
            vals = [
                v["gmv"], v["ing"], v["pago"], v["personal"], v["transportes"],
                v["marketing"], v["hosting"], v["gestoria"], v["seguros"],
                v["telefonia"], v["suministros"], v["otros"], v["amort"], v["fin"],
            ]
            for i, val in enumerate(vals, 3):
                ws.cell(r, i).value = round(val, 2)
                money_fmt(ws.cell(r, i))
            ws.cell(r, 17).value = f"=SUM(E{r}:P{r})"
            ws.cell(r, 18).value = f"=D{r}-Q{r}"
            ws.cell(r, 19).value = f"=MAX(0,R{r}*{IS})"
            ws.cell(r, 20).value = f"=R{r}-S{r}"
            for c in range(17, 21):
                money_fmt(ws.cell(r, c))
            r += 1

    for c in range(1, 21):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["D"].width = 18


def write_resumen_block(ws, start_row: int, scenario: str, label: str, fill: PatternFill):
    ws.cell(start_row, 1).value = label
    ws.cell(start_row, 1).font = Font(bold=True)
    ws.cell(start_row, 1).fill = fill
    headers = ["Año", "GMV", "Ingresos 15%", "Total gastos", "RAI", "IS", "Neto", "Margen neto"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(start_row + 1, i, h)
        style_header(cell)
    for i, t in enumerate(annual_totals(scenario)):
        rr = start_row + 2 + i
        ws.cell(rr, 1).value = t["year"]
        for c, key in enumerate(["gmv", "ing", "gast", "rai", "tax", "neto"], 2):
            ws.cell(rr, c).value = round(t[key], 2)
            money_fmt(ws.cell(rr, c))
            ws.cell(rr, c).fill = fill
        ws.cell(rr, 1).fill = fill
        ws.cell(rr, 8).value = (t["neto"] / t["ing"]) if t["ing"] else 0
        pct_fmt(ws.cell(rr, 8))
        ws.cell(rr, 8).fill = fill


def ensure_sheet(wb, name: str, index: int | None = None):
    if name in wb.sheetnames:
        del wb[name]
    if index is None:
        return wb.create_sheet(name)
    return wb.create_sheet(name, index)


def clear_sheet(ws):
    """Vacía hoja incluyendo celdas combinadas."""
    while ws.merged_cells.ranges:
        ws.unmerge_cells(str(next(iter(ws.merged_cells.ranges))))
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()


def write_kpi_dashboard(wb):
    """Panel KPI + tablas de datos con fórmulas (gráficos dinámicos al abrir Excel)."""
    from openpyxl.chart import BarChart, LineChart, Reference, PieChart
    from openpyxl.chart.label import DataLabelList

    # --- KPI_Datos (fuente de los gráficos; fórmulas → se recalcula al abrir) ---
    ws_d = ensure_sheet(wb, "KPI_Datos", 0)
    clear_sheet(ws_d)
    ws_d["A1"] = "Tablas KPI (no borrar: alimentan los gráficos del Dashboard)"
    ws_d["A1"].font = Font(bold=True, color="2F5233")

    # A) Anual Conservador vía SUMIF sobre PyG Conservador
    ws_d["A3"] = "A. Evolución anual — Conservador (fórmulas dinámicas)"
    ws_d["A3"].font = Font(bold=True)
    for i, h in enumerate(
        ["Año", "GMV", "Ingresos", "Gastos", "Neto", "Margen neto", "Take rate", "Cobertura gastos"],
        1,
    ):
        style_header(ws_d.cell(4, i, h))

    pyg = "'PyG Conservador'"
    for i, year in enumerate(range(1, 6)):
        r = 5 + i
        ws_d.cell(r, 1).value = year
        # Filas PyG: cabecera en fila 2; datos desde fila 3
        ws_d.cell(r, 2).value = f"=SUMIF({pyg}!$A:$A,A{r},{pyg}!$C:$C)"
        ws_d.cell(r, 3).value = f"=SUMIF({pyg}!$A:$A,A{r},{pyg}!$D:$D)"
        ws_d.cell(r, 4).value = f"=SUMIF({pyg}!$A:$A,A{r},{pyg}!$Q:$Q)"
        ws_d.cell(r, 5).value = f"=SUMIF({pyg}!$A:$A,A{r},{pyg}!$T:$T)"
        ws_d.cell(r, 6).value = f"=IF(C{r}=0,0,E{r}/C{r})"
        ws_d.cell(r, 7).value = f"=IF(B{r}=0,0,C{r}/B{r})"
        ws_d.cell(r, 8).value = f"=IF(D{r}=0,0,C{r}/D{r})"
        for c in range(2, 6):
            money_fmt(ws_d.cell(r, c))
        for c in range(6, 9):
            pct_fmt(ws_d.cell(r, c))

    # B) Comparativa escenarios (valores calculados; actualizar con script)
    ws_d["A12"] = "B. Comparativa GMV por escenario"
    ws_d["A12"].font = Font(bold=True)
    for i, h in enumerate(["Año", "Conservador", "Realista", "Optimista"], 1):
        style_header(ws_d.cell(13, i, h))
    for i, year in enumerate(range(1, 6)):
        r = 14 + i
        ws_d.cell(r, 1).value = year
        ws_d.cell(r, 2).value = round(GMV["conservador"][year], 0)
        ws_d.cell(r, 3).value = round(GMV["realista"][year], 0)
        ws_d.cell(r, 4).value = round(GMV["optimista"][year], 0)
        for c in range(2, 5):
            money_fmt(ws_d.cell(r, c))

    ws_d["A21"] = "C. Comparativa Resultado neto por escenario"
    ws_d["A21"].font = Font(bold=True)
    for i, h in enumerate(["Año", "Conservador", "Realista", "Optimista"], 1):
        style_header(ws_d.cell(22, i, h))
    for i, year in enumerate(range(1, 6)):
        r = 23 + i
        ws_d.cell(r, 1).value = year
        for j, sc in enumerate(("conservador", "realista", "optimista"), 2):
            ws_d.cell(r, j).value = round(annual_totals(sc)[year - 1]["neto"], 0)
            money_fmt(ws_d.cell(r, j))

    # D) Estacionalidad Y2 Conservador (filas 15-26 de PyG = año 2 si fila3=Y1 ene)
    # Año1: filas 3-14, Año2: 15-26
    ws_d["A30"] = "D. Estacionalidad GMV — Año 2 Conservador (fórmulas)"
    ws_d["A30"].font = Font(bold=True)
    for i, h in enumerate(["Mes", "GMV", "Ingresos"], 1):
        style_header(ws_d.cell(31, i, h))
    for i, month in enumerate(MONTHS):
        r = 32 + i
        pyg_row = 15 + i  # año 2
        ws_d.cell(r, 1).value = month
        ws_d.cell(r, 2).value = f"={pyg}!C{pyg_row}"
        ws_d.cell(r, 3).value = f"={pyg}!D{pyg_row}"
        money_fmt(ws_d.cell(r, 2))
        money_fmt(ws_d.cell(r, 3))

    # E) Estructura de gastos Y3 (suma categorías)
    ws_d["A46"] = "E. Estructura de gastos — Año 3 Conservador (fórmulas)"
    ws_d["A46"].font = Font(bold=True)
    style_header(ws_d.cell(47, 1, "Partida"))
    style_header(ws_d.cell(47, 2, "Importe"))
    # Año 3 filas PyG: 27-38
    cats = [
        ("Medios de pago", "E"),
        ("Personal", "F"),
        ("Transportes", "G"),
        ("Marketing", "H"),
        ("Hosting/software", "I"),
        ("Gestoría", "J"),
        ("Seguros", "K"),
        ("Telefonía", "L"),
        ("Suministros", "M"),
        ("Otros", "N"),
        ("Amortizaciones", "O"),
        ("Gastos financieros", "P"),
    ]
    for i, (name, col) in enumerate(cats):
        r = 48 + i
        ws_d.cell(r, 1).value = name
        ws_d.cell(r, 2).value = f"=SUMIF({pyg}!$A:$A,3,{pyg}!${col}:${col})"
        money_fmt(ws_d.cell(r, 2))

    # F) KPIs tarjeta en J–K (NO solapar con D–H de la tabla anual; evita circularidad)
    ws_d["J3"] = "F. Tarjetas KPI (Conservador)"
    ws_d["J3"].font = Font(bold=True)
    cards = [
        (5, "GMV Y3", "=B7"),
        (6, "Ingresos Y3", "=C7"),
        (7, "Neto Y3", "=E7"),
        (8, "Margen neto Y3", "=F7"),
        (9, "Take rate", "=G7"),
        (10, "Cobertura gastos Y3", "=H7"),
        (11, "GMV equilibrio mes", round((150 + 550 + 250 + 120 + 30 + 35 + 40 + 25 + 10 + AMORT_ANUAL / 12) / COMISION, 0)),
        (12, "Vendedores obj. Y3", 11),
        (13, "Comisión marketplace", COMISION),
        (14, "Inv. + 15% dividendos", INVERSION_REF * (1 + DIVIDENDO_OBJ_PCT)),
    ]
    style_header(ws_d.cell(4, 10, "KPI"))
    style_header(ws_d.cell(4, 11, "Valor"))
    for r, label, val in cards:
        ws_d.cell(r, 10).value = label
        ws_d.cell(r, 11).value = val
        if label.startswith("Margen") or label.startswith("Take") or label.startswith("Comisión") or label.startswith("Cobertura"):
            pct_fmt(ws_d.cell(r, 11))
        elif isinstance(val, (int, float)) or (isinstance(val, str) and val.startswith("=")):
            money_fmt(ws_d.cell(r, 11))

    for c, w in enumerate([22, 14, 14, 14, 14, 12, 12, 14, 3, 24, 14], 1):
        ws_d.column_dimensions[get_column_letter(c)].width = w

    # --- Dashboard visual ---
    ws = ensure_sheet(wb, "00_KPI_Dashboard", 0)
    clear_sheet(ws)
    ws["A1"] = "DASHBOARD KPI MARKETPLACE — Escenario Conservador (gráficos dinámicos)"
    ws["A1"].font = Font(bold=True, size=14, color="2F5233")
    ws.merge_cells("A1:L1")
    ws["A2"] = (
        "Los gráficos leen KPI_Datos. Si modificas PyG Conservador, recalcula Excel (F9) y los gráficos se actualizan. "
        "GMV ≠ ingreso SL (ingreso = 15% comisión)."
    )
    ws.merge_cells("A2:L2")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 32

    # Tarjetas visibles
    ws["A4"] = "KPIs clave (vinculados)"
    ws["A4"].font = Font(bold=True)
    for i, (label, ref) in enumerate(
        [
            ("GMV Y3", "=KPI_Datos!B7"),
            ("Ingresos Y3", "=KPI_Datos!C7"),
            ("Neto Y3", "=KPI_Datos!E7"),
            ("Margen Y3", "=KPI_Datos!F7"),
            ("Take rate", "=KPI_Datos!G7"),
            ("Cobertura gastos Y3", "=KPI_Datos!H7"),
        ],
        1,
    ):
        c = i
        style_header(ws.cell(5, c, label))
        ws.cell(6, c).value = ref
        if "Margen" in label or "Take" in label or "Cobertura" in label:
            pct_fmt(ws.cell(6, c))
        else:
            money_fmt(ws.cell(6, c))
        ws.cell(6, c).font = Font(bold=True, size=12)

    # Chart 1: GMV vs Ingresos
    c1 = LineChart()
    c1.title = "1. GMV vs Ingresos comisión (Conservador)"
    c1.style = 10
    c1.y_axis.title = "€"
    c1.x_axis.title = "Año"
    c1.height = 10
    c1.width = 15
    data = Reference(ws_d, min_col=2, min_row=4, max_col=3, max_row=9)
    cats = Reference(ws_d, min_col=1, min_row=5, max_row=9)
    c1.add_data(data, titles_from_data=True)
    c1.set_categories(cats)
    ws.add_chart(c1, "A8")

    # Chart 2: Neto
    c2 = BarChart()
    c2.type = "col"
    c2.title = "2. Resultado neto anual (Conservador)"
    c2.style = 10
    c2.y_axis.title = "€"
    c2.height = 10
    c2.width = 12
    data2 = Reference(ws_d, min_col=5, min_row=4, max_row=9)
    c2.add_data(data2, titles_from_data=True)
    c2.set_categories(cats)
    ws.add_chart(c2, "I8")

    # Chart 3: Escenarios GMV
    c3 = BarChart()
    c3.type = "col"
    c3.grouping = "clustered"
    c3.title = "3. GMV por escenario (C / R / O)"
    c3.style = 10
    c3.height = 10
    c3.width = 15
    data3 = Reference(ws_d, min_col=2, min_row=13, max_col=4, max_row=18)
    cats3 = Reference(ws_d, min_col=1, min_row=14, max_row=18)
    c3.add_data(data3, titles_from_data=True)
    c3.set_categories(cats3)
    ws.add_chart(c3, "A25")

    # Chart 4: Neto escenarios
    c4 = LineChart()
    c4.title = "4. Resultado neto por escenario"
    c4.style = 12
    c4.height = 10
    c4.width = 15
    data4 = Reference(ws_d, min_col=2, min_row=22, max_col=4, max_row=27)
    cats4 = Reference(ws_d, min_col=1, min_row=23, max_row=27)
    c4.add_data(data4, titles_from_data=True)
    c4.set_categories(cats4)
    ws.add_chart(c4, "I25")

    # Chart 5: Estacionalidad
    c5 = BarChart()
    c5.type = "col"
    c5.title = "5. Estacionalidad GMV Año 2 (Conservador)"
    c5.style = 10
    c5.height = 10
    c5.width = 18
    data5 = Reference(ws_d, min_col=2, min_row=31, max_row=43)
    cats5 = Reference(ws_d, min_col=1, min_row=32, max_row=43)
    c5.add_data(data5, titles_from_data=True)
    c5.set_categories(cats5)
    ws.add_chart(c5, "A42")

    # Chart 6: Pie gastos Y3
    c6 = PieChart()
    c6.title = "6. Estructura de gastos Año 3 (Conservador)"
    labels6 = Reference(ws_d, min_col=1, min_row=48, max_row=59)
    data6 = Reference(ws_d, min_col=2, min_row=47, max_row=59)
    c6.add_data(data6, titles_from_data=True)
    c6.set_categories(labels6)
    c6.dataLabels = DataLabelList()
    c6.dataLabels.showPercent = True
    c6.dataLabels.showVal = False
    c6.dataLabels.showCatName = False
    c6.height = 12
    c6.width = 14
    ws.add_chart(c6, "I42")

    ws["A58"] = "Leyenda de KPIs utilizados"
    ws["A58"].font = Font(bold=True)
    legend = [
        "GMV: volumen de ventas de vendedores (actividad; no es facturación propia de la SL).",
        "Ingresos: comisión 15% sobre GMV (ingreso de la SL).",
        "Take rate: Ingresos / GMV (debe ser ~15%).",
        "Cobertura de gastos: Ingresos / Gastos (>100% = ingresos cubren gastos).",
        "Margen neto: Resultado neto / Ingresos.",
        "Estacionalidad: refleja pico turístico estival (alineado a EOTR Zamora).",
        "Escenarios C/R/O: Conservador = referencia; Realista/Optimista = sensibilidad.",
    ]
    for i, t in enumerate(legend):
        ws.cell(59 + i, 1).value = t
        ws.merge_cells(start_row=59 + i, start_column=1, end_row=59 + i, end_column=12)

    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # Activar dashboard al abrir
    wb.active = ws


def build():
    wb = openpyxl.load_workbook(SRC)

    # --- PyG Conservador (hoja principal renombrada/reescrita) ---
    ws_c = wb[wb.sheetnames[0]]
    # Si ya existe 00_KPI_Dashboard como primera, localizar PyG
    if ws_c.title.startswith("00_") or ws_c.title.startswith("KPI"):
        for name in wb.sheetnames:
            if "Conservador" in name or name == "PyG mensual":
                ws_c = wb[name]
                break
    ws_c.title = "PyG Conservador"
    clear_sheet(ws_c)
    write_pyg(
        ws_c,
        "conservador",
        "ESCENARIO CONSERVADOR (activo) — GMV≠ingreso SL | Y1: venta jul–dic | Amortización desde lanzamiento",
    )

    ws_r = ensure_sheet(wb, "PyG Realista", 1)
    write_pyg(
        ws_r,
        "realista",
        "ESCENARIO REALISTA (hipótesis) — mayor GMV y gasto; no validado con ventas reales",
    )

    ws_o = ensure_sheet(wb, "PyG Optimista", 2)
    write_pyg(
        ws_o,
        "optimista",
        "ESCENARIO OPTIMISTA (hipótesis) — escalado agresivo; no validado con ventas reales",
    )

    # --- Parámetros ---
    ws_par = None
    for name in wb.sheetnames:
        if name.startswith("Par"):
            ws_par = wb[name]
            break
    if ws_par is None:
        ws_par = ensure_sheet(wb, "Parametros", 3)
    clear_sheet(ws_par)
    ws_par["A1"] = "Parámetro"
    ws_par["B1"] = "Valor"
    ws_par["C1"] = "Comentario"
    for col in "ABC":
        style_header(ws_par[f"{col}1"])
    params = [
        (2, "% comisión", COMISION, "Ingresos SL = GMV × 15%"),
        (3, "IVA referencia", 0.21, "No entra en PyG operativa"),
        (4, "IS orientativo", IS, "No es liquidación fiscal"),
        (5, "Escenario de referencia", "CONSERVADOR", "PyG Conservador = base memoria/ICECYL"),
        (6, "GMV Y1 conservador", GMV["conservador"][1], f"→ {GMV['conservador'][1]*COMISION:,.0f} € comisión (6 meses venta)"),
        (7, "GMV Y2 conservador", GMV["conservador"][2], f"→ {GMV['conservador'][2]*COMISION:,.0f} € comisión"),
        (8, "GMV Y3 conservador", GMV["conservador"][3], f"→ {GMV['conservador'][3]*COMISION:,.0f} € comisión"),
        (9, "% medios de pago / GMV", PCT_PAGO, "Pasarela/Connect orientativo"),
        (10, "Amortización anual", AMORT_ANUAL, "30.000/5; en Y1 solo jul–dic"),
        (11, "Inversión referencia", INVERSION_REF, "Activo tecnológico / umbral ICECYL"),
        (12, "Objetivo dividendos 5 años", DIVIDENDO_OBJ_PCT, "≥15% sobre la inversión (además de recuperarla)"),
        (13, "Neto acumulado mínimo", INVERSION_REF * (1 + DIVIDENDO_OBJ_PCT), "Recuperar 30k + 4,5k dividendos"),
        (14, "Compras mercancía", 0, "Intermediación: sin COGS"),
        (15, "Suavizado Y1", "Sí", "Amort. y marketing comercial desde lanzamiento; H1 costes mínimos"),
    ]
    for r, a, b, c in params:
        ws_par.cell(r, 1).value = a
        ws_par.cell(r, 2).value = b
        ws_par.cell(r, 3).value = c
        if isinstance(b, float) and b <= 1:
            pct_fmt(ws_par.cell(r, 2))
        elif isinstance(b, (int, float)):
            money_fmt(ws_par.cell(r, 2))
    ws_par.column_dimensions["A"].width = 28
    ws_par.column_dimensions["B"].width = 16
    ws_par.column_dimensions["C"].width = 55

    # --- Resumen anual (3 bloques) ---
    for name in list(wb.sheetnames):
        if name.startswith("Resumen"):
            ws_res = wb[name]
            break
    else:
        ws_res = ensure_sheet(wb, "Resumen anual")

    clear_sheet(ws_res)

    ws_res["A1"] = "Resumen anual — tres escenarios (ingresos = GMV × 15%)"
    ws_res["A1"].font = Font(bold=True, size=13, color="2F5233")
    ws_res.merge_cells("A1:H1")

    write_resumen_block(ws_res, 3, "conservador", "🔴 Conservador (referencia)", PatternFill("solid", fgColor="F8D7DA"))
    write_resumen_block(ws_res, 12, "realista", "🟡 Realista (hipótesis)", PatternFill("solid", fgColor="FFF3CD"))
    write_resumen_block(ws_res, 21, "optimista", "🟢 Optimista (hipótesis)", PatternFill("solid", fgColor="D4EDDA"))

    ws_res["A30"] = (
        "Y1 suavizado en conservador: amortización solo desde julio (capitalización en lanzamiento); "
        "marketing comercial en H2; H1 = estructura mínima. "
        "Detalle mes a mes en hojas PyG Conservador / Realista / Optimista."
    )
    ws_res.merge_cells("A30:H30")
    ws_res["A30"].alignment = Alignment(wrap_text=True)
    ws_res.row_dimensions[30].height = 40
    for c in range(1, 9):
        ws_res.column_dimensions[get_column_letter(c)].width = 14

    # --- Drivers ---
    for name in list(wb.sheetnames):
        if name.startswith("Drivers"):
            ws_d = wb[name]
            break
    else:
        ws_d = ensure_sheet(wb, "Drivers y equilibrio")

    clear_sheet(ws_d)

    ws_d["A1"] = "Drivers (conservador) y punto de equilibrio"
    ws_d["A1"].font = Font(bold=True, size=13, color="2F5233")
    drivers = [
        # Año, vendedores, GMV€/vend/mes, meses venta, GMV anual, ingresos 15%
        (1, 5, 1_333.33, 6, 40_000, 6_000),
        (2, 8, 1_458.33, 12, 140_000, 21_000),
        (3, 11, 1_666.67, 12, 220_000, 33_000),
        (4, 13, 1_794.87, 12, 280_000, 42_000),
        (5, 15, 2_000.00, 12, 360_000, 54_000),
    ]
    hdr = ["Año", "Vendedores", "GMV €/vend/mes", "Meses venta", "GMV anual", "Ingresos 15%"]
    for i, h in enumerate(hdr, 1):
        style_header(ws_d.cell(3, i, h))
    for i, row in enumerate(drivers):
        for c, v in enumerate(row, 1):
            ws_d.cell(4 + i, c).value = v
            if c in (3, 5, 6):
                money_fmt(ws_d.cell(4 + i, c))

    # equilibrio Y3+ (opex conservador contenido + amortización)
    fixed = 150 + 550 + 250 + 120 + 30 + 35 + 40 + 25 + 10 + (AMORT_ANUAL / 12)
    ws_d["A10"] = "Umbral orientativo Y3+ (con amortización; opex contenido)"
    ws_d["A10"].font = Font(bold=True)
    ws_d["A11"] = "Gastos fijos mensuales ≈"
    ws_d["B11"] = round(fixed, 2)
    money_fmt(ws_d["B11"])
    ws_d["A12"] = "GMV mínimo mensual (÷15%)"
    ws_d["B12"] = round(fixed / COMISION, 2)
    money_fmt(ws_d["B12"])
    ws_d["A13"] = "Con 11 vendedores → GMV/vend/mes"
    ws_d["B13"] = round((fixed / COMISION) / 11, 2)
    money_fmt(ws_d["B13"])
    ws_d["A15"] = "Objetivo financiero (conservador)"
    ws_d["A15"].font = Font(bold=True)
    ws_d["A16"] = "Inversión de referencia"
    ws_d["B16"] = INVERSION_REF
    money_fmt(ws_d["B16"])
    ws_d["A17"] = "Dividendos mínimos (15%)"
    ws_d["B17"] = INVERSION_REF * DIVIDENDO_OBJ_PCT
    money_fmt(ws_d["B17"])
    ws_d["A18"] = "Neto acumulado objetivo (5 años)"
    ws_d["B18"] = INVERSION_REF * (1 + DIVIDENDO_OBJ_PCT)
    money_fmt(ws_d["B18"])
    cons = annual_totals("conservador")
    cum = sum(t["neto"] for t in cons)
    ws_d["A19"] = "Neto acumulado proyectado (conservador)"
    ws_d["B19"] = round(cum, 0)
    money_fmt(ws_d["B19"])
    ws_d["C19"] = "OK" if cum >= INVERSION_REF * (1 + DIVIDENDO_OBJ_PCT) else "REVISAR"
    for c, w in enumerate([38, 14, 16, 12, 12, 14], 1):
        ws_d.column_dimensions[get_column_letter(c)].width = w

    # --- Escenarios hoja resumen compacta ---
    for name in list(wb.sheetnames):
        if "Escenario" in name or "v" in name.lower() and "as" in name.lower():
            if name.startswith("Escenarios") or "3" in name:
                ws_e = wb[name]
                break
    else:
        ws_e = ensure_sheet(wb, "Escenarios 3 vías")

    # find sheet
    ws_e = None
    for name in wb.sheetnames:
        if name.startswith("Escenarios"):
            ws_e = wb[name]
            break
    if ws_e is None:
        ws_e = ensure_sheet(wb, "Escenarios 3 vías")

    clear_sheet(ws_e)

    ws_e["A1"] = "Comparativa rápida (detalle mensual en PyG Conservador / Realista / Optimista)"
    ws_e["A1"].font = Font(bold=True, size=12, color="2F5233")
    ws_e.merge_cells("A1:F1")
    for i, h in enumerate(["Escenario", "Año", "GMV", "Ingresos 15%", "Gastos", "RAI"], 1):
        style_header(ws_e.cell(3, i, h))
    fills = {
        "conservador": PatternFill("solid", fgColor="F8D7DA"),
        "realista": PatternFill("solid", fgColor="FFF3CD"),
        "optimista": PatternFill("solid", fgColor="D4EDDA"),
    }
    labels = {"conservador": "Conservador", "realista": "Realista", "optimista": "Optimista"}
    rr = 4
    for sc in ("conservador", "realista", "optimista"):
        for t in annual_totals(sc):
            ws_e.cell(rr, 1).value = labels[sc]
            ws_e.cell(rr, 2).value = t["year"]
            ws_e.cell(rr, 3).value = round(t["gmv"], 0)
            ws_e.cell(rr, 4).value = round(t["ing"], 0)
            ws_e.cell(rr, 5).value = round(t["gast"], 0)
            ws_e.cell(rr, 6).value = round(t["rai"], 0)
            for c in range(1, 7):
                ws_e.cell(rr, c).fill = fills[sc]
            for c in range(3, 7):
                money_fmt(ws_e.cell(rr, c))
            rr += 1
    for c, w in enumerate([14, 8, 14, 14, 14, 14], 1):
        ws_e.column_dimensions[get_column_letter(c)].width = w

    # --- Notas ---
    ws_n = None
    for name in wb.sheetnames:
        if name.startswith("Notas"):
            ws_n = wb[name]
            break
    if ws_n:
        clear_sheet(ws_n)
        notes = [
            ("GMV", "No es ingreso de la SL."),
            ("Ingresos", "15% comisión sobre GMV."),
            ("Y1 suavizado", "Amortización y marketing comercial desde julio; H1 costes mínimos de estructura."),
            ("3 PyG", "Conservador = referencia. Realista/Optimista = sensibilidad mensual completa."),
            ("Compras", "0 € — sin reventa de stock."),
            ("IS/IVA", "IS orientativo; IVA fuera de PyG."),
            ("Subvenciones", "No incluidas como ingreso de explotación."),
            ("Dashboard", "Abrir hoja 00_KPI_Dashboard: gráficos dinámicos desde KPI_Datos / PyG."),
        ]
        ws_n["A1"] = "Tema"
        ws_n["B1"] = "Criterio"
        style_header(ws_n["A1"])
        style_header(ws_n["B1"])
        for i, (a, b) in enumerate(notes, 2):
            ws_n.cell(i, 1).value = a
            ws_n.cell(i, 2).value = b

    # Dashboard KPI + gráficos (al final, para referenciar PyG ya escrito)
    # Eliminar dashboards previos si existen
    for name in list(wb.sheetnames):
        if name.startswith("00_KPI") or name == "KPI_Datos":
            del wb[name]
    write_kpi_dashboard(wb)

    wb.save(OUT)
    print(f"Saved {OUT}")
    print("\n=== TOTALES ===")
    for sc in ("conservador", "realista", "optimista"):
        print(f"\n-- {sc.upper()} --")
        for t in annual_totals(sc):
            print(
                f"Y{t['year']}: GMV={t['gmv']:,.0f} Ing={t['ing']:,.0f} "
                f"Gast={t['gast']:,.0f} RAI={t['rai']:,.0f} Neto={t['neto']:,.0f}"
            )


if __name__ == "__main__":
    build()
